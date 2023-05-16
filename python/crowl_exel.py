import openpyxl

# 딕셔너리
my_dict = {
    'name': {'bong': {'zening': {"ff": 2, "gg": 3,"gg": 3,"gg": 3,"gg": 3}}},
    'age': {'bong': {'zening': {"ff": 2, "gg": 3}}},
    'gender': {'bong': {'zening': {"ff": 2, "gg": 3}}}
}

# 엑셀 파일 생성
wb = openpyxl.Workbook()
ws = wb.active

# 딕셔너리에서 값 추출하여 엑셀에 저장
row_num = 1
for key, value in my_dict.items():
    for subkey, subvalue in value.items():
        for subsubkey, subsubvalue in subvalue.items():
            ws.cell(row=row_num, column=1, value=key)
            ws.cell(row=row_num, column=2, value=subkey)
            ws.cell(row=row_num, column=3, value=subsubkey)
            ws.cell(row=row_num, column=4, value=subsubvalue['ff'])
            ws.cell(row=row_num, column=5, value=subsubvalue['gg'])
            row_num += 1

# 엑셀 파일 저장
wb.save('test.xlsx')